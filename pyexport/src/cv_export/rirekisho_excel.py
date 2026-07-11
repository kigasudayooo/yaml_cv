"""data.yaml（既存の履歴書データ形式）から履歴書 .xlsx を生成する。"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from cv_export.styling import JAPANESE_FONT, THIN_BORDER, load_rirekisho_data

TITLE_FONT = Font(name=JAPANESE_FONT, size=16, bold=True)
LABEL_FONT = Font(name=JAPANESE_FONT, size=8)
VALUE_FONT = Font(name=JAPANESE_FONT, size=11)
HEADING_FONT = Font(name=JAPANESE_FONT, size=10, bold=True)

# 1行に収める項目（ふりがな・氏名・電話番号など）は折り返さない。
# 折り返すと行の高さを超えた分が下のセルに重なって表示されてしまうため。
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
# 趣味・志望動機など複数行に渡る自由記述欄のみ折り返しを許可する。
WRAP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _set(ws: Worksheet, cell: str, value: object, font: Font, align: Alignment = LEFT) -> None:
    c = ws[cell]
    c.value = value
    c.font = font
    c.alignment = align


def _history_rows(items: list[dict[str, Any]]) -> list[str]:
    lines = []
    for item in items:
        year = str(item.get("year", ""))
        month = str(item.get("month", ""))
        value = str(item.get("value", ""))
        prefix = f"{year}年{month}月" if year or month else ""
        lines.append(f"{prefix} {value}".strip())
    return lines


def _border_range(ws: Worksheet, cell_range: str) -> None:
    """指定範囲のすべてのセルに罫線を引く（実際に内容がある区画のみに使う）。"""
    for row in ws[cell_range]:
        for cell in row:
            cell.border = THIN_BORDER


def build_rirekisho_workbook(data: dict[str, Any]) -> Workbook:
    """履歴書データから openpyxl Workbook を組み立てる。

    紙の履歴書用紙に近い見た目になるよう、実際に内容がある区画にのみ
    罫線を引く（空の余白行にまで罫線を引くと方眼紙のように見えてしまうため）。

    Args:
        data: load_rirekisho_data で読み込んだ data.yaml 相当の dict。

    Returns:
        構築済みの openpyxl Workbook。
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "履歴書"
    ws.sheet_view.showGridLines = False

    # A-F列のみ使用。ラベル列は狭め、値列(結合セルを含む)は住所等が
    # 折り返さずに収まるよう広めに確保する。
    for col, width in zip("ABCDEF", [11, 17, 11, 17, 11, 17], strict=True):
        ws.column_dimensions[col].width = width
    ws.sheet_format.defaultRowHeight = 20

    border_ranges: list[str] = []

    _set(ws, "A1", "履歴書", TITLE_FONT, CENTER)
    ws.merge_cells("A1:C2")
    _set(ws, "D1", str(data.get("date", "")), VALUE_FONT)
    ws.merge_cells("D1:F1")
    border_ranges.append("A1:F2")

    row = 4
    basic_start = row
    _set(ws, f"A{row}", "ふりがな", LABEL_FONT)
    _set(ws, f"B{row}", str(data.get("name_kana", "")), VALUE_FONT)
    ws.merge_cells(f"B{row}:F{row}")
    row += 1
    _set(ws, f"A{row}", "氏名", LABEL_FONT)
    _set(ws, f"B{row}", str(data.get("name", "")), Font(name=JAPANESE_FONT, size=14, bold=True))
    ws.merge_cells(f"B{row}:F{row}")
    ws.row_dimensions[row].height = 26
    row += 1
    _set(ws, f"A{row}", "生年月日", LABEL_FONT)
    _set(ws, f"B{row}", str(data.get("birth_day", "")), VALUE_FONT)
    ws.merge_cells(f"B{row}:D{row}")
    _set(ws, f"E{row}", "性別", LABEL_FONT)
    _set(ws, f"F{row}", str(data.get("gender", "")), VALUE_FONT)
    row += 1
    _set(ws, f"A{row}", "携帯電話", LABEL_FONT)
    _set(ws, f"B{row}", str(data.get("cell_phone", "")), VALUE_FONT)
    ws.merge_cells(f"B{row}:C{row}")
    _set(ws, f"D{row}", "E-MAIL", LABEL_FONT)
    _set(ws, f"E{row}", str(data.get("email", "")), VALUE_FONT)
    ws.merge_cells(f"E{row}:F{row}")
    border_ranges.append(f"A{basic_start}:F{row}")
    row += 2

    for kana_key, addr_label, zip_key, addr_key, tel_key, fax_key in (
        ("address_kana", "現住所", "address_zip", "address", "tel", "fax"),
        ("address_kana2", "連絡先", "address_zip2", "address2", "tel2", "fax2"),
    ):
        addr_start = row
        _set(ws, f"A{row}", "ふりがな", LABEL_FONT)
        _set(ws, f"B{row}", str(data.get(kana_key, "")), VALUE_FONT)
        ws.merge_cells(f"B{row}:F{row}")
        row += 1
        _set(ws, f"A{row}", addr_label, LABEL_FONT)
        zip_ = str(data.get(zip_key, ""))
        addr = str(data.get(addr_key, ""))
        _set(ws, f"B{row}", f"〒{zip_}　{addr}" if zip_ or addr else "", VALUE_FONT)
        ws.merge_cells(f"B{row}:F{row}")
        row += 1
        _set(ws, f"A{row}", "電話", LABEL_FONT)
        _set(ws, f"B{row}", str(data.get(tel_key, "")), VALUE_FONT)
        _set(ws, f"C{row}", "FAX", LABEL_FONT)
        _set(ws, f"D{row}", str(data.get(fax_key, "")), VALUE_FONT)
        border_ranges.append(f"A{addr_start}:F{row}")
        row += 2

    history_start = row
    _set(ws, f"A{row}", "学歴・職歴", HEADING_FONT)
    row += 1
    for line in _history_rows(data.get("education", [])) or []:
        _set(ws, f"A{row}", line, VALUE_FONT)
        ws.merge_cells(f"A{row}:F{row}")
        row += 1
    _set(ws, f"A{row}", "職歴", HEADING_FONT)
    row += 1
    for line in _history_rows(data.get("experience", [])) or []:
        _set(ws, f"A{row}", line, VALUE_FONT)
        ws.merge_cells(f"A{row}:F{row}")
        row += 1
    _set(ws, f"A{row}", "以上", VALUE_FONT)
    border_ranges.append(f"A{history_start}:F{row}")
    row += 2

    licences = _history_rows(data.get("licences", [])) or []
    if licences:
        lic_start = row
        _set(ws, f"A{row}", "免許・資格", HEADING_FONT)
        row += 1
        for line in licences:
            _set(ws, f"A{row}", line, VALUE_FONT)
            ws.merge_cells(f"A{row}:F{row}")
            row += 1
        border_ranges.append(f"A{lic_start}:F{row - 1}")
        row += 1

    _set(ws, f"A{row}", "通勤時間", LABEL_FONT)
    _set(ws, f"B{row}", str(data.get("commuting_time", "")), VALUE_FONT)
    _set(ws, f"C{row}", "扶養家族", LABEL_FONT)
    _set(ws, f"D{row}", str(data.get("dependents", "")), VALUE_FONT)
    _set(ws, f"E{row}", "配偶者", LABEL_FONT)
    _set(ws, f"F{row}", str(data.get("spouse", "")), VALUE_FONT)
    border_ranges.append(f"A{row}:F{row}")
    row += 2

    for label, key in (
        ("趣味・特技", "hobby"),
        ("志望動機", "motivation"),
        ("本人希望記入欄", "request"),
    ):
        _set(ws, f"A{row}", label, HEADING_FONT)
        row += 1
        box_start = row
        _set(ws, f"A{row}", str(data.get(key, "")), VALUE_FONT, align=WRAP_LEFT)
        ws.merge_cells(f"A{row}:F{row + 2}")
        for r in range(row, row + 3):
            ws.row_dimensions[r].height = 20
        border_ranges.append(f"A{box_start}:F{row + 2}")
        row += 4

    for cell_range in border_ranges:
        _border_range(ws, cell_range)

    return wb


def generate_rirekisho_excel(input_file: Path, output_file: Path) -> None:
    """data.yaml から履歴書 .xlsx を生成してファイルに保存する。

    Args:
        input_file: data.yaml へのパス。
        output_file: 出力する .xlsx ファイルパス。
    """
    data = load_rirekisho_data(input_file)
    wb = build_rirekisho_workbook(data)
    wb.save(str(output_file))
